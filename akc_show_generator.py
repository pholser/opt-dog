"""
akc_show_generator.py
=====================
Generates synthetic AKC all-breed conformation show datasets in the
workbook format defined by the show scheduling data model.

Usage
-----
    # Default mid-sized show
    python akc_show_generator.py

    # Custom show with explicit parameters
    python akc_show_generator.py \\
        --output my_show.xlsx \\
        --seed 99 \\
        --size large \\
        --rings 12 \\
        --group-rings 2 \\
        --conflict-opt-in-rate 0.20

    # All options
    python akc_show_generator.py --help

Parameters
----------
--output            Output file path (default: akc_show_synthetic.xlsx)
--seed              Random seed for reproducibility (default: 42)
--size              Show size preset: small | medium | large (default: medium)
                      small:  ~100 breeds, fewer entries per breed
                      medium: ~189 breeds (all AKC breeds), moderate entries
                      large:  ~189 breeds, higher entries per breed
--rings             Total number of rings (default: 10)
--group-rings       Number of rings designated for Group/BIS judging (default: 2)
--conflict-opt-in-rate
                    Fraction of handlers who opt in for conflict protection
                    (default: 0.15)
--show-date         Show date in YYYY-MM-DD format (default: 2025-09-20)
--club-name         Club name (default: Bluegrass Kennel Club, Inc.)
--venue-name        Venue name (default: Kentucky Exposition Center)
--venue-address     Venue address
--judging-start     Judging start time HH:MM (default: 08:00)
--lunch-start       Lunch window start HH:MM (default: 11:30)
--lunch-end         Lunch window end HH:MM (default: 13:30)
--lunch-duration    Lunch break duration in minutes (default: 45)
--slot-minutes      Time slot size in minutes (default: 10)
"""

import argparse
import math
import random
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# AKC Reference Data
# All 189 AKC-recognized breeds and varieties as of 2025, with group and
# equipment type assignments.
# ---------------------------------------------------------------------------

GROUPS = [
    ("G1", "Sporting"),
    ("G2", "Hound"),
    ("G3", "Working"),
    ("G4", "Terrier"),
    ("G5", "Toy"),
    ("G6", "Non-Sporting"),
    ("G7", "Herding"),
]

# (breed_name, variety, group_id, equipment_type)
# equipment_type: table | ramp | floor
ALL_BREEDS = [
    # --- Sporting ---
    ("Brittanys",                              None,           "G1", "floor"),
    ("Pointers",                               None,           "G1", "floor"),
    ("Pointers (German Shorthaired)",          None,           "G1", "floor"),
    ("Pointers (German Wirehaired)",           None,           "G1", "floor"),
    ("Retrievers (Chesapeake Bay)",            None,           "G1", "floor"),
    ("Retrievers (Curly-Coated)",              None,           "G1", "floor"),
    ("Retrievers (Flat-Coated)",               None,           "G1", "floor"),
    ("Retrievers (Golden)",                    None,           "G1", "floor"),
    ("Retrievers (Labrador)",                  None,           "G1", "floor"),
    ("Retrievers (Nova Scotia Duck Tolling)",  None,           "G1", "floor"),
    ("Setters (English)",                      None,           "G1", "floor"),
    ("Setters (Gordon)",                       None,           "G1", "floor"),
    ("Setters (Irish)",                        None,           "G1", "floor"),
    ("Setters (Irish Red and White)",          None,           "G1", "floor"),
    ("Spaniels (American Water)",              None,           "G1", "floor"),
    ("Spaniels (Boykin)",                      None,           "G1", "floor"),
    ("Spaniels (Clumber)",                     None,           "G1", "ramp"),
    ("Spaniels (Cocker)",                      "Black",        "G1", "floor"),
    ("Spaniels (Cocker)",                      "A.S.C.O.B.",  "G1", "floor"),
    ("Spaniels (Cocker)",                      "Parti-Color",  "G1", "floor"),
    ("Spaniels (English Cocker)",              None,           "G1", "floor"),
    ("Spaniels (English Springer)",            None,           "G1", "floor"),
    ("Spaniels (Field)",                       None,           "G1", "floor"),
    ("Spaniels (Irish Water)",                 None,           "G1", "floor"),
    ("Spaniels (Sussex)",                      None,           "G1", "ramp"),
    ("Spaniels (Welsh Springer)",              None,           "G1", "floor"),
    ("Vizslas",                                None,           "G1", "floor"),
    ("Weimaraners",                            None,           "G1", "floor"),
    ("Wirehaired Pointing Griffons",           None,           "G1", "floor"),
    ("Wirehaired Vizslas",                     None,           "G1", "floor"),
    # --- Hound ---
    ("Afghan Hounds",                          None,           "G2", "floor"),
    ("Basset Hounds",                          None,           "G2", "ramp"),
    ("Beagles",                                "13 Inch",      "G2", "floor"),
    ("Beagles",                                "15 Inch",      "G2", "floor"),
    ("Black and Tan Coonhounds",               None,           "G2", "floor"),
    ("Bloodhounds",                            None,           "G2", "floor"),
    ("Bluetick Coonhounds",                    None,           "G2", "floor"),
    ("Borzois",                                None,           "G2", "floor"),
    ("Cirnechi dell'Etna",                     None,           "G2", "ramp"),
    ("Dachshunds",                             "Longhaired",   "G2", "table"),
    ("Dachshunds",                             "Smooth",       "G2", "table"),
    ("Dachshunds",                             "Wirehaired",   "G2", "table"),
    ("Grand Basset Griffon Vendeens",          None,           "G2", "floor"),
    ("Greyhounds",                             None,           "G2", "floor"),
    ("Harriers",                               None,           "G2", "floor"),
    ("Ibizan Hounds",                          None,           "G2", "floor"),
    ("Irish Wolfhounds",                       None,           "G2", "floor"),
    ("Norwegian Elkhounds",                    None,           "G2", "floor"),
    ("Otterhounds",                            None,           "G2", "floor"),
    ("Petits Bassets Griffons Vendeens",       None,           "G2", "floor"),
    ("Pharaoh Hounds",                         None,           "G2", "floor"),
    ("Plotts",                                 None,           "G2", "floor"),
    ("Redbone Coonhounds",                     None,           "G2", "floor"),
    ("Rhodesian Ridgebacks",                   None,           "G2", "floor"),
    ("Salukis",                                None,           "G2", "floor"),
    ("Scottish Deerhounds",                    None,           "G2", "floor"),
    ("Sloughis",                               None,           "G2", "floor"),
    ("Treeing Walker Coonhounds",              None,           "G2", "floor"),
    ("Whippets",                               None,           "G2", "floor"),
    # --- Working ---
    ("Akitas",                                 None,           "G3", "floor"),
    ("Alaskan Malamutes",                      None,           "G3", "floor"),
    ("Anatolian Shepherd Dogs",                None,           "G3", "floor"),
    ("Bernese Mountain Dogs",                  None,           "G3", "floor"),
    ("Black Russian Terriers",                 None,           "G3", "floor"),
    ("Boerboels",                              None,           "G3", "floor"),
    ("Boxers",                                 None,           "G3", "floor"),
    ("Bullmastiffs",                           None,           "G3", "floor"),
    ("Cane Corsos",                            None,           "G3", "floor"),
    ("Doberman Pinschers",                     None,           "G3", "floor"),
    ("Dogo Argentinos",                        None,           "G3", "floor"),
    ("Dogues de Bordeaux",                     None,           "G3", "floor"),
    ("German Pinschers",                       None,           "G3", "floor"),
    ("Giant Schnauzers",                       None,           "G3", "floor"),
    ("Great Danes",                            None,           "G3", "floor"),
    ("Great Pyrenees",                         None,           "G3", "floor"),
    ("Greater Swiss Mountain Dogs",            None,           "G3", "floor"),
    ("Komondorok",                             None,           "G3", "floor"),
    ("Kuvaszok",                               None,           "G3", "floor"),
    ("Leonbergers",                            None,           "G3", "floor"),
    ("Mastiffs",                               None,           "G3", "floor"),
    ("Newfoundlands",                          None,           "G3", "floor"),
    ("Portuguese Water Dogs",                  None,           "G3", "floor"),
    ("Rottweilers",                            None,           "G3", "floor"),
    ("Saint Bernards",                         None,           "G3", "floor"),
    ("Samoyeds",                               None,           "G3", "floor"),
    ("Siberian Huskies",                       None,           "G3", "floor"),
    ("Standard Schnauzers",                    None,           "G3", "floor"),
    ("Tibetan Mastiffs",                       None,           "G3", "floor"),
    # --- Terrier ---
    ("Airedale Terriers",                      None,           "G4", "floor"),
    ("American Hairless Terriers",             None,           "G4", "table"),
    ("American Staffordshire Terriers",        None,           "G4", "floor"),
    ("Australian Terriers",                    None,           "G4", "table"),
    ("Bedlington Terriers",                    None,           "G4", "floor"),
    ("Border Terriers",                        None,           "G4", "table"),
    ("Bull Terriers",                          "Colored",      "G4", "floor"),
    ("Bull Terriers",                          "White",        "G4", "floor"),
    ("Cairn Terriers",                         None,           "G4", "table"),
    ("Cesky Terriers",                         None,           "G4", "table"),
    ("Dandie Dinmont Terriers",                None,           "G4", "table"),
    ("Fox Terriers (Smooth)",                  None,           "G4", "floor"),
    ("Fox Terriers (Wire)",                    None,           "G4", "floor"),
    ("Glen of Imaal Terriers",                 None,           "G4", "table"),
    ("Irish Terriers",                         None,           "G4", "floor"),
    ("Kerry Blue Terriers",                    None,           "G4", "ramp"),
    ("Lakeland Terriers",                      None,           "G4", "floor"),
    ("Manchester Terriers (Standard)",         None,           "G4", "table"),
    ("Miniature Bull Terriers",                None,           "G4", "floor"),
    ("Miniature Schnauzers",                   None,           "G4", "floor"),
    ("Norfolk Terriers",                       None,           "G4", "table"),
    ("Norwich Terriers",                       None,           "G4", "table"),
    ("Parson Russell Terriers",                None,           "G4", "table"),
    ("Rat Terriers",                           None,           "G4", "table"),
    ("Russell Terriers",                       None,           "G4", "table"),
    ("Scottish Terriers",                      None,           "G4", "table"),
    ("Sealyham Terriers",                      None,           "G4", "table"),
    ("Skye Terriers",                          None,           "G4", "table"),
    ("Soft Coated Wheaten Terriers",           None,           "G4", "floor"),
    ("Staffordshire Bull Terriers",            None,           "G4", "floor"),
    ("Welsh Terriers",                         None,           "G4", "floor"),
    ("West Highland White Terriers",           None,           "G4", "floor"),
    # --- Toy ---
    ("Affenpinschers",                         None,           "G5", "table"),
    ("Brussels Griffons",                      None,           "G5", "table"),
    ("Cavalier King Charles Spaniels",         None,           "G5", "table"),
    ("Chihuahuas",                             "Long Coat",    "G5", "table"),
    ("Chihuahuas",                             "Smooth Coat",  "G5", "table"),
    ("Chinese Cresteds",                       None,           "G5", "table"),
    ("English Toy Spaniels",                   "B & PC",       "G5", "table"),
    ("English Toy Spaniels",                   "KC & R",       "G5", "table"),
    ("Havanese",                               None,           "G5", "table"),
    ("Italian Greyhounds",                     None,           "G5", "table"),
    ("Japanese Chin",                          None,           "G5", "table"),
    ("Maltese",                                None,           "G5", "table"),
    ("Manchester Terriers (Toy)",              None,           "G5", "table"),
    ("Miniature Pinschers",                    None,           "G5", "table"),
    ("Papillons",                              None,           "G5", "table"),
    ("Pekingese",                              None,           "G5", "table"),
    ("Pomeranians",                            None,           "G5", "table"),
    ("Poodles (Toy)",                          None,           "G5", "table"),
    ("Pugs",                                   None,           "G5", "table"),
    ("Russian Toys",                           None,           "G5", "table"),
    ("Shih Tzu",                               None,           "G5", "table"),
    ("Silky Terriers",                         None,           "G5", "table"),
    ("Toy Fox Terriers",                       None,           "G5", "table"),
    ("Yorkshire Terriers",                     None,           "G5", "table"),
    # --- Non-Sporting ---
    ("American Eskimo Dogs",                   None,           "G6", "floor"),
    ("Bichons Frises",                         None,           "G6", "floor"),
    ("Boston Terriers",                        None,           "G6", "floor"),
    ("Bulldogs",                               None,           "G6", "ramp"),
    ("Chinese Shar-Pei",                       None,           "G6", "ramp"),
    ("Chow Chows",                             None,           "G6", "ramp"),
    ("Cotons de Tulear",                       None,           "G6", "table"),
    ("Dalmatians",                             None,           "G6", "floor"),
    ("Finnish Spitz",                          None,           "G6", "floor"),
    ("French Bulldogs",                        None,           "G6", "floor"),
    ("Keeshonden",                             None,           "G6", "ramp"),
    ("Lhasa Apsos",                            None,           "G6", "table"),
    ("Lowchen",                                None,           "G6", "table"),
    ("Norwegian Lundehunds",                   None,           "G6", "floor"),
    ("Poodles (Miniature)",                    None,           "G6", "floor"),
    ("Poodles (Standard)",                     None,           "G6", "floor"),
    ("Schipperkes",                            None,           "G6", "table"),
    ("Shiba Inu",                              None,           "G6", "floor"),
    ("Tibetan Spaniels",                       None,           "G6", "table"),
    ("Tibetan Terriers",                       None,           "G6", "floor"),
    ("Xoloitzcuintli",                         None,           "G6", "floor"),
    # --- Herding ---
    ("Australian Cattle Dogs",                 None,           "G7", "floor"),
    ("Australian Shepherds",                   None,           "G7", "floor"),
    ("Beaucerons",                             None,           "G7", "floor"),
    ("Bearded Collies",                        None,           "G7", "floor"),
    ("Belgian Laekenois",                      None,           "G7", "floor"),
    ("Belgian Malinois",                       None,           "G7", "floor"),
    ("Belgian Sheepdogs",                      None,           "G7", "floor"),
    ("Belgian Tervuren",                       None,           "G7", "floor"),
    ("Bergamasco Sheepdogs",                   None,           "G7", "floor"),
    ("Berger Picards",                         None,           "G7", "floor"),
    ("Border Collies",                         None,           "G7", "floor"),
    ("Bouviers des Flandres",                  None,           "G7", "floor"),
    ("Briards",                                None,           "G7", "floor"),
    ("Cardigan Welsh Corgis",                  None,           "G7", "floor"),
    ("Collies",                                "Rough",        "G7", "floor"),
    ("Collies",                                "Smooth",       "G7", "floor"),
    ("German Shepherd Dogs",                   None,           "G7", "floor"),
    ("Icelandic Sheepdogs",                    None,           "G7", "floor"),
    ("Miniature American Shepherds",           None,           "G7", "floor"),
    ("Norwegian Buhunds",                      None,           "G7", "floor"),
    ("Old English Sheepdogs",                  None,           "G7", "floor"),
    ("Polish Lowland Sheepdogs",               None,           "G7", "floor"),
    ("Pulik",                                  None,           "G7", "floor"),
    ("Pumik",                                  None,           "G7", "floor"),
    ("Pyrenean Shepherds",                     None,           "G7", "floor"),
    ("Shetland Sheepdogs",                     None,           "G7", "floor"),
    ("Swedish Vallhunds",                      None,           "G7", "floor"),
]

# Breeds considered "popular" (higher entry counts)
POPULAR_BREEDS = {
    "Retrievers (Labrador)", "Retrievers (Golden)", "French Bulldogs",
    "German Shepherd Dogs", "Poodles (Standard)", "Australian Shepherds",
    "Bulldogs", "Boxers", "Doberman Pinschers", "Border Collies",
    "Shetland Sheepdogs", "Beagles", "Miniature Schnauzers",
    "Rhodesian Ridgebacks", "Whippets", "Belgian Tervuren",
    "Dachshunds", "Siberian Huskies", "Boston Terriers",
    "West Highland White Terriers", "Pointers (German Shorthaired)",
    "Weimaraners", "Vizslas", "Spaniels (English Springer)",
}

# Moderate popularity
MODERATE_BREEDS = {
    "Brittanys", "Spaniels (Welsh Springer)", "Setters (Irish)",
    "Setters (English)", "Basset Hounds", "Greyhounds", "Irish Wolfhounds",
    "Borzois", "Akitas", "Alaskan Malamutes", "Bernese Mountain Dogs",
    "Rottweilers", "Samoyeds", "Portuguese Water Dogs", "Great Danes",
    "Newfoundlands", "Airedale Terriers", "Cairn Terriers",
    "Kerry Blue Terriers", "Soft Coated Wheaten Terriers",
    "Norwich Terriers", "Border Terriers", "Pugs", "Chihuahuas",
    "Yorkshire Terriers", "Shih Tzu", "Havanese", "Pomeranians",
    "Cavalier King Charles Spaniels", "Bichons Frises", "Chow Chows",
    "Keeshonden", "Dalmatians", "Shiba Inu", "Tibetan Spaniels",
    "Australian Cattle Dogs", "Cardigan Welsh Corgis", "Pumik",
    "Miniature American Shepherds", "Beaucerons", "Bouviers des Flandres",
    "Belgian Malinois", "Briards", "Icelandic Sheepdogs",
}

# Judge name pool
JUDGE_NAME_POOL = [
    "Mrs. Sarah Callahan",    "Mr. Robert Tanner",       "Ms. Patricia Voss",
    "Mr. Gerald Hutchins",    "Mrs. Diana Farrell",       "Mr. Thomas Wexler",
    "Ms. Carol Beaumont",     "Mr. James Okafor",         "Mrs. Linda Ng",
    "Mr. Andrew Marsh",       "Mrs. Cynthia Holt",        "Mr. Daniel Pryce",
    "Ms. Karen Bautista",     "Mr. Paul Oduya",           "Mrs. Frances Doyle",
    "Mr. Victor Ashworth",    "Ms. Helen Moreau",         "Mr. Ian Blackwell",
    "Mrs. Janet Kimura",      "Mr. Lawrence Petrov",      "Ms. Meredith Osei",
    "Mr. Nathan Szabo",       "Mrs. Olivia Reyes",        "Mr. Patrick Yuen",
    "Ms. Ruth Engstrom",      "Mr. Sebastian Kowalczyk",  "Mrs. Teresa Nakagawa",
    "Mr. Ulrich Brandt",      "Ms. Valentina Escobar",    "Mr. William Achebe",
    "Mrs. Xiulan Hoffmann",   "Mr. Yusuf Abramowitz",     "Ms. Zelda Figueiredo",
    "Mr. Adrian Beausoleil",  "Mrs. Brigitte Oyelaran",   "Mr. Carlo Lindstrom",
]

# Handler name pool
HANDLER_NAME_POOL = [
    "Alice Morton",       "Bob Keller",        "Carol Singh",       "David Chen",
    "Eve Larson",         "Frank Russo",       "Grace Kim",         "Henry Patel",
    "Irene Walsh",        "Jack Torres",       "Karen Noble",       "Leo Fernandez",
    "Mia Johansson",      "Nick Osei",         "Olivia Grant",      "Paul Dubois",
    "Quinn Ramirez",      "Rosa Bennett",      "Sam Yong",          "Tina Kowalski",
    "Uma Petrov",         "Victor Nair",       "Wendy Schulz",      "Xander Brooks",
    "Yuki Tanaka",        "Zoe Andersen",      "Aaron Levy",        "Beth Nakamura",
    "Chris Muller",       "Dana Reyes",        "Ethan Clarke",      "Fiona Okonkwo",
    "George Hoffman",     "Hannah Wu",         "Ian Byrne",         "Julia Vasquez",
    "Kyle Nguyen",        "Laura Eriksson",    "Marco De Luca",     "Nina Park",
    "Oscar Lindqvist",    "Penny Adeyemi",     "Ralph Zimmerman",   "Sandra Ito",
    "Tim Blackwood",      "Uma Christensen",   "Vince Abbate",      "Willa Strand",
    "Xavier Peralta",     "Yvonne Mwangi",
]


# ---------------------------------------------------------------------------
# Entry count generation
# ---------------------------------------------------------------------------

# Entry ranges by size preset and popularity tier
# (class_dog_range, class_bitch_range, specials_dog_range, specials_bitch_range)
ENTRY_RANGES = {
    #         popular                   moderate                  rare
    "small":  ((2,8),  (2,9),  (1,3), (1,3),
               (1,4),  (2,5),  (0,2), (0,2),
               (0,2),  (1,3),  (0,1), (0,1)),
    "medium": ((5,14), (6,16), (2,6), (2,6),
               (2,7),  (2,8),  (1,4), (1,4),
               (0,3),  (1,4),  (0,2), (0,2)),
    "large":  ((8,22), (10,25),(3,8), (3,8),
               (3,10), (4,11), (2,6), (2,6),
               (1,4),  (1,5),  (0,2), (0,2)),
}

def gen_entries(breed_name, size_preset, rng):
    """Generate realistic entry counts for a breed."""
    ranges = ENTRY_RANGES[size_preset]
    if breed_name in POPULAR_BREEDS:
        cd_r, cb_r, sd_r, sb_r = ranges[0], ranges[1], ranges[2], ranges[3]
    elif breed_name in MODERATE_BREEDS:
        cd_r, cb_r, sd_r, sb_r = ranges[4], ranges[5], ranges[6], ranges[7]
    else:
        cd_r, cb_r, sd_r, sb_r = ranges[8], ranges[9], ranges[10], ranges[11]

    cd = rng.randint(*cd_r)
    cb = rng.randint(*cb_r)
    sd = rng.randint(*sd_r)
    sb = rng.randint(*sb_r)
    nr = rng.randint(0, 1) if size_preset != "small" else 0

    # Ensure at least 2 total entries
    while cd + cb + sd + sb < 2:
        cb += 1

    # nonregular_position: most shows run NR after specials; ~20% before
    nr_pos = "before_specials" if (nr > 0 and rng.random() < 0.20) else "after_specials"

    return cd, cb, sd, sb, nr, nr_pos


# ---------------------------------------------------------------------------
# Judge assignment
# ---------------------------------------------------------------------------

# Group judge cap: leave headroom since group judging doesn't count toward
# the 175-entry breed limit but consumes the judge's time.
GROUP_JUDGE_BREED_CAP = 150
OVERFLOW_JUDGE_CAP    = 168
PERMIT_JUDGE_CAP      =  50


def assign_judges(breeds_with_entries, group_judge_map, overflow_judge_ids,
                  permit_judge_id):
    """
    Assign a judge to each breed respecting entry caps.

    Parameters
    ----------
    breeds_with_entries : list of (bname, variety, gid, equip, total_entries)
    group_judge_map     : dict {group_id: judge_id}
    overflow_judge_ids  : list of non-group judge IDs (standard rate)
    permit_judge_id     : judge ID of the permit judge (lower cap)

    Returns
    -------
    dict {(bname, variety): judge_id}
    """
    caps = {jid: GROUP_JUDGE_BREED_CAP for jid in group_judge_map.values()}
    for jid in overflow_judge_ids:
        caps[jid] = OVERFLOW_JUDGE_CAP
    if permit_judge_id:
        caps[permit_judge_id] = PERMIT_JUDGE_CAP

    used   = {jid: 0 for jid in caps}
    result = {}

    overflow_standard = list(overflow_judge_ids)
    all_overflow      = overflow_standard + ([permit_judge_id] if permit_judge_id else [])

    HARD_CAP = 175

    def pick_overflow(n):
        candidates = [(j, caps[j] - used[j]) for j in all_overflow
                      if caps[j] - used[j] >= n]
        if candidates:
            return max(candidates, key=lambda x: x[1])[0]
        # Soft cap exhausted — allow up to the hard 175 limit before
        # falling back to the least-loaded judge unconditionally.
        hard_candidates = [(j, HARD_CAP - used[j]) for j in all_overflow
                           if HARD_CAP - used[j] >= n]
        if hard_candidates:
            return max(hard_candidates, key=lambda x: x[1])[0]
        return min(overflow_standard, key=lambda j: used[j])

    # Group breeds by group, process largest-first within each group
    group_breeds = defaultdict(list)
    for item in breeds_with_entries:
        group_breeds[item[2]].append(item)

    for gid, items in group_breeds.items():
        gj = group_judge_map[gid]
        items_sorted = sorted(items, key=lambda x: x[4], reverse=True)
        for bname, variety, _, equip, total in items_sorted:
            key = (bname, variety)
            if used[gj] + total <= caps[gj]:
                result[key] = gj
                used[gj]   += total
            else:
                assigned     = pick_overflow(total)
                result[key]  = assigned
                used[assigned] += total

    return result, used


# ---------------------------------------------------------------------------
# Workbook styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
_thin       = Side(style="thin", color="BBBBBB")
CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def write_sheet(wb, title, headers, rows, first=False):
    """Create and style a worksheet."""
    ws = wb.active if first else wb.create_sheet(title)
    ws.title = title

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 28

    for r_idx, row in enumerate(rows, start=2):
        ws.append([v if v is not None else "" for v in row])
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for cell in ws[r_idx]:
            cell.font   = DATA_FONT
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(max_len + 2, 10), 42)
        )

    return ws


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_show(
    output_path        = "akc_show_synthetic.xlsx",
    seed               = 42,
    size               = "medium",
    n_rings            = 10,
    n_group_rings      = 2,
    conflict_opt_in    = 0.15,
    show_date          = "2025-09-20",
    club_name          = "Bluegrass Kennel Club, Inc.",
    venue_name         = "Kentucky Exposition Center",
    venue_address      = "937 Phillips Lane, Louisville, KY 40209",
    judging_start      = "08:00",
    lunch_start        = "11:30",
    lunch_end          = "13:30",
    lunch_duration_min = 45,
    slot_minutes       = 10,
):
    """
    Generate a synthetic AKC all-breed show dataset.

    Parameters match the CLI arguments described in the module docstring.
    Returns the path of the written workbook.
    """
    rng = random.Random(seed)

    show_id  = "SHOW001"
    group_ids = [g[0] for g in GROUPS]

    # ------------------------------------------------------------------
    # Select breed list based on size preset
    # ------------------------------------------------------------------
    if size == "small":
        # Sample ~60% of breeds from each group
        group_breed_lists = defaultdict(list)
        for b in ALL_BREEDS:
            group_breed_lists[b[2]].append(b)
        breeds_raw = []
        for gid in group_ids:
            pool = group_breed_lists[gid]
            k    = max(3, math.ceil(len(pool) * 0.60))
            breeds_raw += rng.sample(pool, min(k, len(pool)))
    else:
        breeds_raw = list(ALL_BREEDS)

    n_breeds = len(breeds_raw)

    # ------------------------------------------------------------------
    # Generate entry counts
    # ------------------------------------------------------------------
    breed_entry_data = {}
    for bname, variety, gid, equip in breeds_raw:
        breed_entry_data[(bname, variety)] = gen_entries(bname, size, rng)  # (cd,cb,sd,sb,nr,nr_pos)

    total_dogs = sum(sum(v[:5]) for v in breed_entry_data.values())  # first 5 are numeric counts

    # ------------------------------------------------------------------
    # Determine judge panel
    # ------------------------------------------------------------------
    # 7 group judges + enough overflow judges to stay within 175/judge cap
    # + 1 permit judge + 1 BIS-only judge
    overflow_capacity = OVERFLOW_JUDGE_CAP
    group_capacity    = GROUP_JUDGE_BREED_CAP * 7
    overflow_needed   = max(0, total_dogs - group_capacity)
    n_overflow_std    = math.ceil(overflow_needed / overflow_capacity) + 1  # +1 buffer

    # Cap at available judge names minus 7 group judges and 2 (permit + BIS)
    max_overflow = len(JUDGE_NAME_POOL) - 9
    n_overflow_std = min(n_overflow_std, max_overflow)

    judge_pool  = list(JUDGE_NAME_POOL)
    rng.shuffle(judge_pool)

    judge_rows       = []   # (judge_id, show_id, name, is_permit, override)
    group_judge_map  = {}   # {group_id: judge_id}
    overflow_ids     = []
    permit_judge_id  = None
    bis_judge_id     = None

    jcounter = 1
    def next_jid():
        nonlocal jcounter
        jid = f"J{jcounter:02d}"
        jcounter += 1
        return jid

    # 7 group judges
    for gid, gname in GROUPS:
        jid  = next_jid()
        name = judge_pool.pop()
        judge_rows.append((jid, show_id, name, False, ""))
        group_judge_map[gid] = jid

    # Overflow standard judges
    for _ in range(n_overflow_std):
        jid  = next_jid()
        name = judge_pool.pop()
        judge_rows.append((jid, show_id, name, False, ""))
        overflow_ids.append(jid)

    # Permit judge (if we have names left)
    if judge_pool:
        jid             = next_jid()
        name            = judge_pool.pop()
        permit_judge_id = jid
        judge_rows.append((jid, show_id, name, True, ""))

    # BIS judge
    jid          = next_jid()
    name         = judge_pool.pop()
    bis_judge_id = jid
    judge_rows.append((jid, show_id, name, False, ""))

    # ------------------------------------------------------------------
    # Assign judges to breeds
    # ------------------------------------------------------------------
    breeds_with_entries = [
        (bname, variety, gid, equip, sum(breed_entry_data[(bname, variety)][:5]))
        for bname, variety, gid, equip in breeds_raw
    ]
    breed_judge_map, judge_used = assign_judges(
        breeds_with_entries, group_judge_map, overflow_ids, permit_judge_id
    )

    # ------------------------------------------------------------------
    # Build breeds and breed_entries lists
    # ------------------------------------------------------------------
    breeds        = []
    breed_entries = []
    for i, (bname, variety, gid, equip) in enumerate(breeds_raw):
        bid   = f"B{i+1:03d}"
        key   = (bname, variety)
        jid   = breed_judge_map[key]
        rate  = "permit" if jid == permit_judge_id else "standard"
        cd, cb, sd, sb, nr, nr_pos = breed_entry_data[key]
        breeds.append((bid, show_id, bname, variety or "", gid, equip, rate))
        breed_entries.append((bid, show_id, cd, cb, sd, sb, nr, nr_pos))

    # bid lookup for downstream use
    bid_map = {(bname, variety or ""): bid
               for bid, _, bname, variety, *_ in breeds}

    # ------------------------------------------------------------------
    # Rings
    # ------------------------------------------------------------------
    rings          = []
    ring_positions = {}
    cols           = math.ceil(n_rings / 2)
    for i in range(1, n_rings + 1):
        rid       = str(i)
        is_group  = i > (n_rings - n_group_rings)
        if is_group:
            w, l = 50.0, 70.0
        elif i <= n_rings // 2:
            w, l = 40.0, 50.0
        else:
            w, l = 32.0, 42.0
        rings.append((rid, show_id, is_group, w, l, ""))
        col = (i - 1) % cols
        row = (i - 1) // cols
        ring_positions[rid] = (col * 55, row * 65)

    ring_distances = []
    ring_ids = [str(i) for i in range(1, n_rings + 1)]
    for i, ra in enumerate(ring_ids):
        for rb in ring_ids[i+1:]:
            ax, ay = ring_positions[ra]
            bx, by = ring_positions[rb]
            dist   = round(((ax - bx)**2 + (ay - by)**2)**0.5, 1)
            ring_distances.append((show_id, ra, rb, dist))

    # ------------------------------------------------------------------
    # Handlers
    # ------------------------------------------------------------------
    # Expand name pool with suffixes if needed
    handler_names = list(HANDLER_NAME_POOL)
    for suffix in ["Jr.", "Sr.", "II", "III"]:
        for name in HANDLER_NAME_POOL:
            handler_names.append(f"{name} {suffix}")

    # Target ~1 handler per 8 dogs, minimum 50
    n_handlers   = max(50, min(len(handler_names), total_dogs // 8))
    handler_names = handler_names[:n_handlers]

    handlers = []
    for i, name in enumerate(handler_names):
        hid    = f"H{i+1:03d}"
        opt_in = rng.random() < conflict_opt_in
        handlers.append((hid, show_id, name, opt_in))

    handler_ids = [h[0] for h in handlers]

    # ------------------------------------------------------------------
    # Dogs
    # ------------------------------------------------------------------
    dogs        = []
    dog_counter = 1
    for (bid, _, bname, variety, gid, equip, rate), (_, __, cd, cb, sd, sb, nr, nr_pos) in \
            zip(breeds, breed_entries):

        def make_dogs(n, entry_type, sex):
            nonlocal dog_counter
            result = []
            for _ in range(n):
                did = f"D{dog_counter:04d}"
                dog_counter += 1
                hid = rng.choice(handler_ids)
                result.append((did, show_id, bid, hid, entry_type, sex, ""))
            return result

        dogs += make_dogs(cd, "class",     "dog")
        dogs += make_dogs(cb, "class",     "bitch")
        dogs += make_dogs(sd, "specials",  "dog")
        dogs += make_dogs(sb, "specials",  "bitch")
        dogs += make_dogs(nr, "nonregular","dog")

    # ------------------------------------------------------------------
    # BreedJudgeAssignments
    # ------------------------------------------------------------------
    bja_rows = []
    for i, (bid, _, bname, variety, gid, *_) in enumerate(breeds):
        key = (bname, variety if variety else None)
        jid = breed_judge_map[key]
        bja_rows.append((f"BJA{i+1:03d}", show_id, bid, jid, True, "", ""))

    # ------------------------------------------------------------------
    # Write workbook
    # ------------------------------------------------------------------
    wb = Workbook()

    write_sheet(wb, "Show", [
        "show_id", "club_name", "venue_name", "venue_address", "show_date",
        "show_hours_start", "show_hours_end", "judging_start",
        "lunch_window_start", "lunch_window_end", "lunch_duration_min",
        "time_slot_minutes", "indoor", "notes"
    ], [(
        show_id, club_name, venue_name, venue_address, show_date,
        "06:00", "19:00", judging_start,
        lunch_start, lunch_end, lunch_duration_min, slot_minutes, True,
        f"The {club_name} assumes no responsibility for personal injury or "
        f"loss of property at the show grounds."
    )], first=True)

    write_sheet(wb, "Rings",
        ["ring_id", "show_id", "is_group_ring", "width_ft", "length_ft", "notes"],
        rings)

    write_sheet(wb, "RingDistances",
        ["show_id", "ring_id_a", "ring_id_b", "distance_ft"],
        ring_distances)

    write_sheet(wb, "Groups",
        ["group_id", "show_id", "group_name"],
        [(gid, show_id, gname) for gid, gname in GROUPS])

    write_sheet(wb, "Breeds",
        ["breed_id", "show_id", "breed_name", "variety",
         "group_id", "equipment_type", "judging_rate"],
        breeds)

    write_sheet(wb, "BreedEntries",
        ["breed_id", "show_id", "n_class_dogs", "n_class_bitches",
         "n_specials_dogs", "n_specials_bitches", "n_nonregular",
         "nonregular_position"],
        breed_entries)

    write_sheet(wb, "Judges",
        ["judge_id", "show_id", "judge_name", "is_permit", "judging_rate_override"],
        judge_rows)

    write_sheet(wb, "BreedJudgeAssignments",
        ["assignment_id", "show_id", "breed_id", "judge_id",
         "is_original", "original_judge_id", "substitution_rule"],
        bja_rows)

    write_sheet(wb, "GroupJudgeAssignments",
        ["show_id", "group_id", "judge_id"],
        [(show_id, gid, jid) for gid, jid in group_judge_map.items()])

    write_sheet(wb, "BISJudgeAssignment",
        ["show_id", "judge_id"],
        [(show_id, bis_judge_id)])

    write_sheet(wb, "Handlers",
        ["handler_id", "show_id", "handler_name", "conflict_opt_in"],
        handlers)

    write_sheet(wb, "Dogs",
        ["dog_id", "show_id", "breed_id", "handler_id",
         "entry_type", "sex", "registered_name"],
        dogs)

    wb.save(output_path)

    # ------------------------------------------------------------------
    # Summary report
    # ------------------------------------------------------------------
    opt_in_count = sum(1 for h in handlers if h[3])
    over_limit   = {jid: v for jid, v in judge_used.items() if v > 175}

    print(f"Show dataset written to: {output_path}")
    print(f"  Size preset     : {size}")
    print(f"  Breeds          : {n_breeds}")
    print(f"  Total dogs      : {len(dogs)}")
    print(f"  Judges          : {len(judge_rows)}  "
          f"({len(group_judge_map)} group + {len(overflow_ids)} overflow "
          f"+ {'1 permit' if permit_judge_id else '0 permit'} + 1 BIS)")
    print(f"  Rings           : {n_rings}  ({n_group_rings} group rings)")
    print(f"  Handlers        : {len(handlers)}  ({opt_in_count} opted in)")
    print()
    print("  Judge entry load:")
    for jid, name, is_permit, _ in [(r[0], r[2], r[3], None) for r in judge_rows]:
        used  = judge_used.get(jid, 0)
        role  = "BIS only" if jid == bis_judge_id else \
                ("permit"  if is_permit            else \
                ("group"   if jid in group_judge_map.values() else "overflow"))
        flag  = "  *** OVER 175 ***" if used > 175 else ""
        print(f"    {jid}  {name:<30s}  {used:>4d} entries  [{role}]{flag}")

    if over_limit:
        print()
        print("  WARNING: Some judges exceed the 175-entry limit.")
        print("  Consider increasing --rings or adjusting --size.")
    else:
        print()
        print("  All judges within 175-entry limit. OK.")

    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Generate a synthetic AKC all-breed show dataset.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--output",             default="akc_show_synthetic.xlsx")
    p.add_argument("--seed",               type=int,   default=42)
    p.add_argument("--size",               choices=["small","medium","large"],
                                           default="medium")
    p.add_argument("--rings",              type=int,   default=10,
                   dest="n_rings")
    p.add_argument("--group-rings",        type=int,   default=2,
                   dest="n_group_rings")
    p.add_argument("--conflict-opt-in-rate", type=float, default=0.15,
                   dest="conflict_opt_in")
    p.add_argument("--show-date",          default="2025-09-20")
    p.add_argument("--club-name",          default="Bluegrass Kennel Club, Inc.")
    p.add_argument("--venue-name",         default="Kentucky Exposition Center")
    p.add_argument("--venue-address",
                   default="937 Phillips Lane, Louisville, KY 40209")
    p.add_argument("--judging-start",      default="08:00")
    p.add_argument("--lunch-start",        default="11:30")
    p.add_argument("--lunch-end",          default="13:30")
    p.add_argument("--lunch-duration",     type=int,   default=45,
                   dest="lunch_duration_min")
    p.add_argument("--slot-minutes",       type=int,   default=10)
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    generate_show(
        output_path        = args.output,
        seed               = args.seed,
        size               = args.size,
        n_rings            = args.n_rings,
        n_group_rings      = args.n_group_rings,
        conflict_opt_in    = args.conflict_opt_in,
        show_date          = args.show_date,
        club_name          = args.club_name,
        venue_name         = args.venue_name,
        venue_address      = args.venue_address,
        judging_start      = args.judging_start,
        lunch_start        = args.lunch_start,
        lunch_end          = args.lunch_end,
        lunch_duration_min = args.lunch_duration_min,
        slot_minutes       = args.slot_minutes,
    )
